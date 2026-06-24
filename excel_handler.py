"""엑셀 양식 생성 및 입고 파일 파싱 모듈"""

import json
import os
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

from data import INGREDIENTS

RECEIVE_LOG_FILE = "receive_log.json"

# 엑셀 양식 컬럼 정의 (순서 고정)
COLUMNS = [
    ("A", "식자재명",     18),
    ("B", "분류",         10),
    ("C", "납품수량",     12),
    ("D", "단위",         8),
    ("E", "단가(원/단위)", 14),
    ("F", "공급가액(원)", 14),
    ("G", "부가세(원)",   12),
    ("H", "합계금액(원)", 14),
    ("I", "비고",         18),
]


def _border(style="thin"):
    s = Side(style=style, color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border():
    thick = Side(style="medium", color="1A4A7C")
    thin  = Side(style="thin",   color="BDBDBD")
    return Border(left=thick, right=thick, top=thin, bottom=thin)


def generate_template(supplier_name: str = "", category: str = "") -> BytesIO:
    """
    공급업체용 입고 엑셀 양식 생성.
    category 지정 시 해당 카테고리 식자재만, 없으면 전체.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "입고명세서"

    # ── 배경색 상수 ──────────────────────────────
    BLUE_DARK   = "1A4A7C"
    BLUE_MID    = "2B6CB0"
    BLUE_LIGHT  = "EBF4FF"
    YELLOW_HEAD = "FFF8E1"
    GRAY_BG     = "F8F9FA"
    WHITE       = "FFFFFF"

    # ── 제목 영역 (행 1~4) ──────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "한씨막국수 — 식자재 입고 명세서"
    title_cell.font = Font(name="맑은 고딕", size=16, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 메타 정보 행 (2~3)
    meta = [
        ("A2", "공급업체명"),  ("B2", supplier_name or ""),
        ("D2", "분류"),        ("E2", category or "전체"),
        ("G2", "납품일자"),    ("H2", datetime.now().strftime("%Y-%m-%d")),
        ("A3", "담당자"),      ("B3", ""),
        ("D3", "연락처"),      ("E3", ""),
        ("G3", "작성일시"),    ("H3", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for cell_ref, val in meta:
        c = ws[cell_ref]
        c.value = val
        is_label = cell_ref[1] in ("A", "D", "G") or cell_ref in ("A2","D2","G2","A3","D3","G3")
        # 라벨 셀인지 값 셀인지 구분
        if val in ("공급업체명","분류","납품일자","담당자","연락처","작성일시"):
            c.font = Font(name="맑은 고딕", size=9, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=BLUE_MID)
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.font = Font(name="맑은 고딕", size=10, color="333333")
            c.fill = PatternFill("solid", fgColor=YELLOW_HEAD)
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1)
        c.border = _border()

    ws.merge_cells("B2:C2"); ws.merge_cells("E2:F2"); ws.merge_cells("H2:I2")
    ws.merge_cells("B3:C3"); ws.merge_cells("E3:F3"); ws.merge_cells("H3:I3")
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # 빈 구분 행
    ws.row_dimensions[4].height = 6
    for col in range(1, 10):
        ws.cell(4, col).fill = PatternFill("solid", fgColor=BLUE_DARK)

    # ── 컬럼 헤더 (행 5) ────────────────────────
    for col_letter, col_name, col_width in COLUMNS:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        c = ws.cell(5, col_idx)
        c.value = col_name
        c.font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_MID)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = _border()
        ws.column_dimensions[col_letter].width = col_width
    ws.row_dimensions[5].height = 30

    # ── 식자재 데이터 행 (행 6~) ────────────────
    ingredients = {
        name: info for name, info in INGREDIENTS.items()
        if not category or info["category"] == category
    }
    # 카테고리 → 식자재명 순 정렬
    sorted_items = sorted(ingredients.items(),
                          key=lambda x: (x[1]["category"], x[0]))

    current_cat = None
    data_start_row = 6
    row = data_start_row

    for name, info in sorted_items:
        # 카테고리 구분 행
        if info["category"] != current_cat:
            current_cat = info["category"]
            ws.merge_cells(f"A{row}:I{row}")
            cat_cell = ws.cell(row, 1)
            cat_cell.value = f"▶ {current_cat}"
            cat_cell.font = Font(name="맑은 고딕", size=9, bold=True,
                                 color=BLUE_DARK)
            cat_cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
            cat_cell.alignment = Alignment(horizontal="left", vertical="center",
                                           indent=1)
            cat_cell.border = _border()
            ws.row_dimensions[row].height = 18
            row += 1

        bg = WHITE if row % 2 == 0 else GRAY_BG

        # A: 식자재명
        c = ws.cell(row, 1, name)
        c.font = Font(name="맑은 고딕", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = _border()

        # B: 분류
        c = ws.cell(row, 2, info["category"])
        c.font = Font(name="맑은 고딕", size=9, color="666666")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()

        # C: 납품수량 (입력칸 — 노란색)
        c = ws.cell(row, 3, 0)
        c.font = Font(name="맑은 고딕", size=11, bold=True, color="C0392B")
        c.fill = PatternFill("solid", fgColor="FFFDE7")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thick_border()
        c.number_format = "#,##0"

        # D: 단위
        c = ws.cell(row, 4, info["unit"])
        c.font = Font(name="맑은 고딕", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()

        # E: 단가 (입력칸 — 노란색)
        unit_price = round(info["cost_per_unit"])
        c = ws.cell(row, 5, unit_price)
        c.font = Font(name="맑은 고딕", size=11, bold=True, color="C0392B")
        c.fill = PatternFill("solid", fgColor="FFFDE7")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _thick_border()
        c.number_format = "#,##0"

        # F: 공급가액 = C*E / 1.1 (자동계산)
        c = ws.cell(row, 6)
        c.value = f"=ROUND(C{row}*E{row}/1.1,0)"
        c.font = Font(name="맑은 고딕", size=10, color="1A4A7C")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _border()
        c.number_format = "#,##0"

        # G: 부가세 = C*E - F (자동계산)
        c = ws.cell(row, 7)
        c.value = f"=C{row}*E{row}-F{row}"
        c.font = Font(name="맑은 고딕", size=10, color="1A4A7C")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _border()
        c.number_format = "#,##0"

        # H: 합계금액 = C*E (자동계산)
        c = ws.cell(row, 8)
        c.value = f"=C{row}*E{row}"
        c.font = Font(name="맑은 고딕", size=10, bold=True, color=BLUE_DARK)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _border()
        c.number_format = "#,##0"

        # I: 비고
        c = ws.cell(row, 9, "")
        c.font = Font(name="맑은 고딕", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border()

        ws.row_dimensions[row].height = 22
        row += 1

    # ── 합계 행 ────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row, 1, "합  계")
    c.font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border()

    for col_idx, formula_col in [(6, "F"), (7, "G"), (8, "H")]:
        c = ws.cell(row, col_idx)
        c.value = f"=SUM({formula_col}{data_start_row}:{formula_col}{row-1})"
        c.font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_DARK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _border()
        c.number_format = "#,##0"

    c = ws.cell(row, 9, "")
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.border = _border()
    ws.row_dimensions[row].height = 26

    # ── 안내 문구 ──────────────────────────────
    note_row = row + 2
    ws.merge_cells(f"A{note_row}:I{note_row}")
    c = ws.cell(note_row, 1)
    c.value = ("※ 노란색(🟡) 셀만 입력하세요: [납품수량] [단가]  "
               "나머지는 자동 계산됩니다.  "
               "납품 후 이 파일을 재고관리 시스템에 업로드해 주세요.")
    c.font = Font(name="맑은 고딕", size=9, italic=True, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── 인쇄 설정 ──────────────────────────────
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.freeze_panes = "A6"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_upload(file_stream) -> dict:
    """
    업로드된 엑셀을 파싱하여 입고 데이터 반환.
    Returns:
      {
        "supplier": str,
        "date": str,
        "items": [{"name", "qty", "unit", "unit_price", "total"}],
        "errors": [str],
      }
    """
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active

    result = {
        "supplier": str(ws["B2"].value or "").strip(),
        "date": str(ws["H2"].value or datetime.now().strftime("%Y-%m-%d")).strip(),
        "items": [],
        "errors": [],
    }

    # 행 6부터 읽기 (카테고리 구분 행 제외)
    known_names = set(INGREDIENTS.keys())
    for row in ws.iter_rows(min_row=6, values_only=True):
        name = str(row[0] or "").strip()
        if not name or name.startswith("▶") or name == "합  계":
            continue
        if name not in known_names:
            continue  # 모르는 식자재는 스킵

        try:
            qty        = float(row[2] or 0)
            unit       = str(row[3] or "").strip()
            unit_price = float(row[4] or 0)
            total      = round(qty * unit_price)
        except (TypeError, ValueError):
            result["errors"].append(f"{name}: 수량 또는 단가 형식 오류")
            continue

        if qty <= 0:
            continue  # 납품수량 0인 항목은 스킵

        result["items"].append({
            "name": name,
            "qty": qty,
            "unit": unit,
            "unit_price": unit_price,
            "total": total,
        })

    if not result["items"]:
        result["errors"].append("납품수량이 입력된 항목이 없습니다.")

    return result


def apply_receive(parsed: dict, memo: str = "") -> dict:
    """
    파싱된 입고 데이터를 재고에 반영하고 이력 저장.
    단가도 최신 납품 단가로 갱신.
    """
    from inventory import load_stock, save_stock

    stock = load_stock()
    applied = []
    skipped = []

    for item in parsed["items"]:
        name = item["name"]
        if name not in stock:
            skipped.append(name)
            continue

        # 재고 증가 (max_stock 초과 허용 — 실제 납품량 우선)
        before = stock[name]["stock"]
        stock[name]["stock"] = before + item["qty"]

        # 단가 갱신
        if item["unit_price"] > 0:
            stock[name]["cost_per_unit"] = item["unit_price"]

        applied.append({**item, "before": before,
                        "after": stock[name]["stock"]})

    save_stock(stock)

    # 입고 이력 저장
    log_entry = {
        "timestamp": datetime.now().isoformat(),
        "supplier": parsed["supplier"],
        "date": parsed["date"],
        "memo": memo,
        "items": applied,
        "total_amount": sum(i["total"] for i in applied),
    }
    logs = []
    if os.path.exists(RECEIVE_LOG_FILE):
        with open(RECEIVE_LOG_FILE, "r", encoding="utf-8") as f:
            logs = json.load(f)
    logs.append(log_entry)
    with open(RECEIVE_LOG_FILE, "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)

    return {"applied": applied, "skipped": skipped,
            "total_amount": log_entry["total_amount"]}


def get_receive_logs(limit: int = 20) -> list[dict]:
    if not os.path.exists(RECEIVE_LOG_FILE):
        return []
    with open(RECEIVE_LOG_FILE, "r", encoding="utf-8") as f:
        logs = json.load(f)
    return list(reversed(logs))[:limit]
